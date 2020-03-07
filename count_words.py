from openpyxl import Workbook, load_workbook
import sys
from textblob import TextBlob


file_name = sys.argv[1]
wb = load_workbook(sys.argv[1])
ws = wb.active
output_file_name = file_name.replace('.xlsx', '_word_count.xlsx')


current_row = 2
col_comments_heading = 13
col_comments = 14

all_words = {}


while True:
    print('Row ', str(current_row), 'Done')
    if ws.cell(row=current_row, column=1).value:
        if ws.cell(row=current_row, column=col_comments).value:
            current_data = ws.cell(row=current_row, column=col_comments).value + ws.cell(row=current_row, column=col_comments_heading).value
            blob = TextBlob(current_data.replace('/', ' ').replace('.', ' '))
            for sen in blob.sentences:
                for w in blob.words:
                    if w in all_words:
                        all_words[w] += 1
                    else:
                        all_words[w] = 1
        current_row += 1
    else:
        print('All done!')
        break
print(all_words)
print('Saving ', file_name)

current_row = 1
wb_out = Workbook()
ws_out = wb_out.active
for key in all_words:
    ws_out.cell(row=current_row, column=1).value = key
    ws_out.cell(row=current_row, column=2).value = all_words[key]
    current_row += 1

wb_out.save(output_file_name)
