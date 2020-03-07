from openpyxl import Workbook, load_workbook
import sys
from textblob import TextBlob


file_name = sys.argv[1]
wb = load_workbook(sys.argv[1])
ws = wb.active


current_row = 1
col_comments_heading = 13
col_comments = 14
col_polarity = 16
col_subjectivity = 17
col_words = 18
col_sentences = 19


ws.cell(row=current_row, column=col_polarity).value = 'polarity'
ws.cell(row=current_row, column=col_subjectivity).value = 'subjectivity'
ws.cell(row=current_row, column=col_words).value = 'total_words'
ws.cell(row=current_row, column=col_sentences).value = 'total_sentences'
current_row += 1


while True:
    print('Row ', str(current_row), 'Done')
    if ws.cell(row=current_row, column=1).value:
        if ws.cell(row=current_row, column=col_comments).value:
            blob = TextBlob(ws.cell(row=current_row, column=col_comments).value + ws.cell(row=current_row, column=col_comments_heading).value.replace('/', ' ').replace('.', ' '))
            total_polarity = 0
            total_subjectivity = 0
            for sen in blob.sentences:
                total_polarity += sen.sentiment.polarity
                total_subjectivity += sen.sentiment.subjectivity
                ws.cell(row=current_row, column=col_polarity).value = total_polarity
                ws.cell(row=current_row, column=col_subjectivity).value = total_subjectivity
                ws.cell(row=current_row, column=col_words).value = len(blob.words)
                ws.cell(row=current_row, column=col_sentences).value = len(blob.sentences)
        current_row += 1
    else:
        print('All done!')
        break

print('Saving ', file_name)
wb.save(file_name)
