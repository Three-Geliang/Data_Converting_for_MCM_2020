from openpyxl import Workbook, load_workbook
import sys
from textblob import TextBlob


file_name = sys.argv[1]
wb = load_workbook(sys.argv[1])
ws = wb.active
file_output = file_name.replace('.xlsx', '_lower_cased.xlsx')


current_row = 2
current_col = 1
while True:
    print('Row', str(current_row), 'Done')
    if ws.cell(row=current_row, column=1).value:
        current_value = ws.cell(row=current_row, column=current_col).value
        print(current_value)
        if current_value:
            if type(current_value) is str:
                ws.cell(row=current_row, column=current_col).value = current_value.lower()
            current_col += 1
        else:
            current_col = 1
            current_row += 1
    else:
        print('All done!')
        break

print('Saving ', file_output)
wb.save(file_output)
